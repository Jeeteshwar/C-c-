from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

# Prompt user for Amazon product link
product_link = input("Enter Amazon product link: ")

# Set up Selenium webdriver for Chrome
driver = webdriver.Chrome('chromedriver')

# Load Amazon product page
driver.get(product_link)

# Wait for reviews to load
WebDriverWait(driver, 10).until(
    EC.presence_of_element_located((By.ID, 'reviews-medley-footer'))
)

# Find all review elements
review_elements = driver.find_elements_by_xpath("//div[@data-hook='review']")

# Create a new Excel workbook and worksheet
wb = Workbook()
ws = wb.active

# Write review data to worksheet
for i, element in enumerate(review_elements):
    # Get review title
    title = element.find_element_by_xpath(".//a[@data-hook='review-title']/span")
    
    # Get review rating
    rating = element.find_element_by_xpath(".//i[@data-hook='review-star-rating']")
    
    # Get review text
    text = element.find_element_by_xpath(".//span[@data-hook='review-body']")
    
    # Write review data to worksheet
    ws.cell(row=i+1, column=1, value=title.text)
    ws.cell(row=i+1, column=2, value=rating.get_attribute('class'))
    ws.cell(row=i+1, column=3, value=text.text)

# Save workbook to file
wb.save('amazon_reviews.xlsx')

# Close webdriver
driver.quit()
